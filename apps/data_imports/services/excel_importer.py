from __future__ import annotations

import hashlib
import logging
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
import unicodedata

from django.core.exceptions import ValidationError
from django.core.validators import validate_email
from django.db.models import Sum
from django.utils import timezone
from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

from apps.core.normalizers import normalize_email, normalize_parcel_code, normalize_phone, normalize_rut_dv, normalize_rut_number
from apps.core.validators import validate_rut
from apps.data_imports.models import ImportIssue, ImportJob, ImportSheetResult, ImportStatus, IssueSeverity
from apps.finance.models import CommonExpenseDebt, PaymentAgreement, ServiceDebt, ServiceType, UnpaidFine
from apps.notes.models import AdministrativeNote, NoteType
from apps.parcels.models import Parcel
from apps.people.models import OwnershipType, ParcelOwnership, ParcelResident, Person, ResidentType
from apps.utilities.models import CutType, ServiceCut, ServiceHistory
from apps.vehicles.models import Vehicle
from apps.works.models import ParcelWorkStatus

logger = logging.getLogger(__name__)


class ImportCancelledError(Exception):
    pass


@dataclass
class Counter:
    rows_read: int = 0
    inserted: int = 0
    updated: int = 0
    skipped: int = 0
    errors: int = 0
    warnings: int = 0


class ExcelMasterImporter:
    SHEET_REQUIREMENTS = {
        'Datos_Propietarios': ['parcela', 'nombre completo', 'rut'],
        'OTROS DUEÑOS': ['parcela'],
        'RESIDENTES': ['parcela', 'residente'],
        'PPU_LOGOS': ['parcela', 'ppu'],
        'Mora GC': ['parcela', 'mora cg uf', 'total pesos'],
        'DESUDAS AyS': ['parcela', 'total deuda'],
        'MORA CONVENIO': ['parcela', 'total mora'],
        'Multas-Convenios impagas': ['parcela', 'empresa', 'saldo monto'],
        'Cortes Vigentes': ['cliente', 'estado'],
        'HISTORICO AYS': ['parcela', 'solicitante', 'descripcion'],
        'ANOTACIONES': ['parcela', 'fecha', 'anotacion'],
        'OBRAS': ['parcela n', 'cortafuego', 'limpieza'],
    }

    def __init__(
        self,
        file_path: str,
        dry_run: bool = False,
        initiated_by=None,
        sheets: list[str] | None = None,
        column_mapping: dict | None = None,
    ):
        self.file_path = Path(file_path)
        self.dry_run = dry_run
        self.initiated_by = initiated_by
        self.sheets_filter = {s.strip() for s in sheets} if sheets else None
        self.column_mapping = self._normalize_column_mapping(column_mapping or {})
        self._cancel_check_every = 25
        self._operations_since_cancel_check = 0

    def _parser_map(self):
        return {
            'Datos_Propietarios': self._parse_datos_propietarios,
            'OTROS DUEÑOS': self._parse_otros_duenos,
            'RESIDENTES': self._parse_residentes,
            'PPU_LOGOS': self._parse_vehiculos,
            'Mora GC': self._parse_mora_gc,
            'DESUDAS AyS': self._parse_deudas_ays,
            'MORA CONVENIO': self._parse_mora_convenio,
            'Multas-Convenios impagas': self._parse_multas,
            'Cortes Vigentes': self._parse_cortes,
            'HISTORICO AYS': self._parse_historico_ays,
            'ANOTACIONES': self._parse_anotaciones,
            'OBRAS': self._parse_obras,
        }

    def inspect_structure(self):
        workbook = load_workbook(self.file_path, data_only=True)
        parser_map = self._parser_map()
        checks = []
        for sheet_name in parser_map.keys():
            if self.sheets_filter and sheet_name not in self.sheets_filter:
                continue
            required_keywords = self.SHEET_REQUIREMENTS.get(sheet_name, [])
            if sheet_name not in workbook.sheetnames:
                checks.append(
                    {
                        'sheet_name': sheet_name,
                        'exists': False,
                        'header_found': False,
                        'required_keywords': required_keywords,
                        'missing_keywords': required_keywords,
                        'header_row': None,
                    }
                )
                continue
            ws = workbook[sheet_name]
            header_row, headers = self._find_header(ws, required_keywords) if required_keywords else (1, {})
            missing = []
            if required_keywords and headers:
                missing = [
                    keyword
                    for keyword in required_keywords
                    if not any(self._norm_header(keyword) == key or self._norm_header(keyword) in key for key in headers.keys())
                ]
            elif required_keywords:
                missing = list(required_keywords)

            checks.append(
                {
                    'sheet_name': sheet_name,
                    'exists': True,
                    'header_found': bool(header_row),
                    'required_keywords': required_keywords,
                    'missing_keywords': missing,
                    'header_row': header_row or None,
                }
            )

        return {
            'available_sheets': workbook.sheetnames,
            'checks': checks,
        }

    def run(self) -> ImportJob:
        if not self.file_path.exists():
            raise FileNotFoundError(f'Archivo no encontrado: {self.file_path}')

        job = ImportJob.objects.create(
            source_file=self.file_path.name,
            source_hash=self._hash_file(self.file_path),
            source_path=str(self.file_path),
            dry_run=self.dry_run,
            status=ImportStatus.RUNNING,
            initiated_by=self.initiated_by,
        )

        parser_map = self._parser_map()

        workbook = load_workbook(self.file_path, data_only=True)
        cancelled = False
        for sheet_name, parser in parser_map.items():
            if self.sheets_filter and sheet_name not in self.sheets_filter:
                continue
            if self._is_cancel_requested(job):
                cancelled = True
                break
            if sheet_name not in workbook.sheetnames:
                self._issue(job, None, IssueSeverity.WARNING, sheet_name, None, None, 'sheet_missing', 'Hoja no encontrada')
                continue

            sheet_result = ImportSheetResult.objects.create(import_job=job, sheet_name=sheet_name, status=ImportStatus.RUNNING)
            counter = Counter()
            ws = workbook[sheet_name]

            try:
                parser(ws, job, sheet_result, counter)
                sheet_result.status = ImportStatus.PARTIAL if counter.errors else ImportStatus.SUCCESS
            except ImportCancelledError:
                cancelled = True
                sheet_result.status = ImportStatus.CANCELLED
                self._issue(
                    job,
                    sheet_result,
                    IssueSeverity.WARNING,
                    sheet_name,
                    None,
                    None,
                    'job_cancelled',
                    'Importación detenida por solicitud de cancelación.',
                )
            except Exception as exc:  # pragma: no cover
                logger.exception('Error importando hoja %s', sheet_name)
                counter.errors += 1
                self._issue(
                    job,
                    sheet_result,
                    IssueSeverity.ERROR,
                    sheet_name,
                    None,
                    None,
                    'sheet_crash',
                    f'Error crítico en hoja: {exc}',
                )
                sheet_result.status = ImportStatus.FAILED

            sheet_result.rows_read = counter.rows_read
            sheet_result.inserted = counter.inserted
            sheet_result.updated = counter.updated
            sheet_result.skipped = counter.skipped
            sheet_result.errors = counter.errors
            sheet_result.warnings = counter.warnings
            sheet_result.summary = (
                f'rows={counter.rows_read}, inserted={counter.inserted}, updated={counter.updated}, '
                f'skipped={counter.skipped}, errors={counter.errors}, warnings={counter.warnings}'
            )
            sheet_result.save()
            if cancelled:
                break

        self._finalize_job(job, cancelled=cancelled)
        return job

    def _normalize_column_mapping(self, value):
        normalized = {}
        if not isinstance(value, dict):
            return normalized
        for sheet_name, aliases in value.items():
            sheet_key = self._norm_header(sheet_name)
            if not isinstance(aliases, dict):
                continue
            normalized[sheet_key] = {}
            for source_alias, mapped_alias in aliases.items():
                source_key = self._norm_header(source_alias)
                if not source_key:
                    continue
                if isinstance(mapped_alias, list):
                    normalized[sheet_key][source_key] = [self._norm_header(item) for item in mapped_alias if self._norm_header(item)]
                else:
                    mapped_key = self._norm_header(mapped_alias)
                    if mapped_key:
                        normalized[sheet_key][source_key] = [mapped_key]
        return normalized

    def _finalize_job(self, job: ImportJob, cancelled: bool = False):
        aggregates = job.sheet_results.aggregate(
            inserted=Sum('inserted'),
            updated=Sum('updated'),
            skipped=Sum('skipped'),
            errors=Sum('errors'),
            warnings=Sum('warnings'),
        )
        job.total_inserted = aggregates['inserted'] or 0
        job.total_updated = aggregates['updated'] or 0
        job.total_skipped = aggregates['skipped'] or 0
        job.total_errors = aggregates['errors'] or 0
        job.total_warnings = aggregates['warnings'] or 0
        job.finished_at = timezone.now()
        cancelled = cancelled or self._is_cancel_requested(job)

        if cancelled:
            job.status = ImportStatus.CANCELLED
            details = dict(job.details or {})
            details['cancel_requested'] = True
            details.setdefault('cancelled_at', timezone.now().isoformat())
            job.details = details
        elif job.total_errors == 0:
            job.status = ImportStatus.SUCCESS
        elif job.total_inserted > 0 or job.total_updated > 0:
            job.status = ImportStatus.PARTIAL
        else:
            job.status = ImportStatus.FAILED

        job.save(
            update_fields=[
                'total_inserted',
                'total_updated',
                'total_skipped',
                'total_errors',
                'total_warnings',
                'finished_at',
                'status',
                'details',
            ]
        )

    def _is_cancel_requested(self, job: ImportJob) -> bool:
        job.refresh_from_db(fields=['status', 'details'])
        details = job.details or {}
        return job.status == ImportStatus.CANCELLED or bool(details.get('cancel_requested'))

    def _raise_if_cancel_requested(self, job: ImportJob, force: bool = False):
        self._operations_since_cancel_check += 1
        if not force and self._operations_since_cancel_check < self._cancel_check_every:
            return
        self._operations_since_cancel_check = 0
        if self._is_cancel_requested(job):
            raise ImportCancelledError()

    def _hash_file(self, path: Path) -> str:
        digest = hashlib.sha256()
        with path.open('rb') as fh:
            for chunk in iter(lambda: fh.read(4096), b''):
                digest.update(chunk)
        return digest.hexdigest()

    def _issue(
        self,
        job: ImportJob,
        sheet_result: ImportSheetResult | None,
        severity: str,
        sheet_name: str,
        row_number: int | None,
        column_name: str | None,
        error_code: str,
        message: str,
        raw_value: str = '',
    ):
        ImportIssue.objects.create(
            import_job=job,
            sheet_result=sheet_result,
            severity=severity,
            sheet_name=sheet_name,
            row_number=row_number,
            column_name=column_name or '',
            error_code=error_code,
            message=message,
            raw_value=(raw_value or '')[:500],
        )

    def _norm_header(self, value) -> str:
        txt = '' if value is None else str(value)
        txt = unicodedata.normalize('NFKD', txt)
        txt = ''.join(ch for ch in txt if not unicodedata.combining(ch))
        txt = txt.replace('\n', ' ').replace('\r', ' ')
        txt = ''.join(ch if ch.isalnum() else ' ' for ch in txt)
        txt = ' '.join(txt.lower().split())
        return txt

    def _find_header(self, ws, required_keywords: list[str], max_rows: int = 20) -> tuple[int, dict[str, int]]:
        required_norm = [self._norm_header(x) for x in required_keywords]
        for row_idx in range(1, max_rows + 1):
            values = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
            header_map = {}
            for idx, val in enumerate(values, start=1):
                norm = self._norm_header(val)
                if norm:
                    header_map[norm] = idx
            if all(any(req in key for key in header_map.keys()) for req in required_norm):
                return row_idx, header_map
        return 0, {}

    def _cell(self, ws, row: int, col_map: dict[str, int], *aliases: str):
        sheet_mapping = self.column_mapping.get(self._norm_header(ws.title), {})
        for alias in aliases:
            norm_alias = self._norm_header(alias)
            lookup_aliases = [norm_alias]
            lookup_aliases.extend(sheet_mapping.get(norm_alias, []))
            for key, idx in col_map.items():
                if any(candidate == key or candidate in key for candidate in lookup_aliases):
                    return ws.cell(row=row, column=idx).value
        return None

    def _to_int(self, value, default=0):
        if value in (None, ''):
            return default
        try:
            return int(Decimal(str(value).replace(',', '.')))
        except (InvalidOperation, ValueError, TypeError):
            return default

    def _to_decimal(self, value, default=Decimal('0')):
        if value in (None, ''):
            return default
        try:
            return Decimal(str(value).replace(' ', '').replace(',', '.'))
        except (InvalidOperation, ValueError, TypeError):
            return default

    def _to_date(self, value):
        if value in (None, ''):
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)):
            try:
                converted = from_excel(value)
                return converted.date() if isinstance(converted, datetime) else converted
            except Exception:
                return None
        if isinstance(value, str):
            raw = value.strip()
            for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d'):
                try:
                    return datetime.strptime(raw, fmt).date()
                except ValueError:
                    continue
        return None

    def _upsert_parcel(self, raw_code: str, counter: Counter, job: ImportJob, sheet_result: ImportSheetResult, row_number: int):
        self._raise_if_cancel_requested(job)
        code = normalize_parcel_code(raw_code)
        if not code:
            counter.warnings += 1
            self._issue(job, sheet_result, IssueSeverity.WARNING, sheet_result.sheet_name, row_number, 'PARCELA', 'invalid_parcel', 'Parcela inválida', str(raw_code))
            return None

        parcel = Parcel.objects.filter(codigo_parcela_key=code).first()
        if parcel:
            return parcel

        if self.dry_run:
            counter.inserted += 1
            return None

        parcel = Parcel.objects.create(codigo_parcela=code)
        counter.inserted += 1
        return parcel

    def _get_or_create_person(self, *, nombre, rut='', dv='', phone1='', phone2='', email='', notes='', counter: Counter, dry_create=True):
        nombre = (nombre or '').strip()
        rut_norm = normalize_rut_number(rut)
        dv_norm = normalize_rut_dv(dv)
        email_norm = normalize_email(email)

        if rut_norm and dv_norm and not validate_rut(rut_norm, dv_norm):
            dv_norm = ''
        if email_norm:
            try:
                validate_email(email_norm)
            except ValidationError:
                email_norm = ''

        person = None
        if rut_norm:
            person = Person.objects.filter(rut_normalizado=rut_norm).first()
        if not person and email_norm:
            person = Person.objects.filter(email=email_norm, nombre_completo__iexact=nombre).first()
        if not person and nombre:
            person = Person.objects.filter(nombre_completo__iexact=nombre, rut_normalizado='').first()

        defaults = {
            'nombre_completo': nombre or 'Sin nombre',
            'rut': rut_norm,
            'rut_dv': dv_norm,
            'telefono_principal': normalize_phone(phone1),
            'telefono_secundario': normalize_phone(phone2),
            'email': email_norm,
            'notas': notes or '',
            'activo': True,
        }

        if person:
            changed = False
            for field, value in defaults.items():
                if value and getattr(person, field) != value:
                    setattr(person, field, value)
                    changed = True
            if changed and not self.dry_run:
                try:
                    person.save()
                    counter.updated += 1
                except ValidationError:
                    counter.warnings += 1
            return person

        if not dry_create:
            return None

        if self.dry_run:
            counter.inserted += 1
            return None

        try:
            person = Person.objects.create(**defaults)
            counter.inserted += 1
            return person
        except ValidationError:
            counter.warnings += 1
            fallback = {**defaults, 'email': '', 'rut_dv': ''}
            try:
                person = Person.objects.create(**fallback)
                counter.inserted += 1
                return person
            except ValidationError:
                counter.errors += 1
                return None

    def _upsert_ownership(self, parcel, person, tipo, counter: Counter):
        if not parcel or not person:
            return
        lookup = {'parcela': parcel, 'persona': person, 'tipo': tipo}
        existing = ParcelOwnership.objects.filter(**lookup, is_deleted=False).first()
        if existing:
            if not existing.is_active and not self.dry_run:
                existing.is_active = True
                existing.save(update_fields=['is_active', 'updated_at'])
                counter.updated += 1
            else:
                counter.skipped += 1
            return
        if self.dry_run:
            counter.inserted += 1
            return
        ParcelOwnership.objects.create(parcela=parcel, persona=person, tipo=tipo, is_active=True)
        counter.inserted += 1

    def _parse_datos_propietarios(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'nombre completo', 'rut'])
        if not header_row:
            counter.errors += 1
            self._issue(job, sheet_result, IssueSeverity.ERROR, ws.title, None, None, 'header_not_found', 'No se encontró encabezado en Datos_Propietarios')
            return

        for row in range(header_row + 1, ws.max_row + 1):
            raw_parcel = self._cell(ws, row, headers, 'parcela')
            raw_name = self._cell(ws, row, headers, 'nombre completo')
            if not raw_parcel and not raw_name:
                continue

            counter.rows_read += 1
            parcel = self._upsert_parcel(raw_parcel, counter, job, sheet_result, row)
            person = self._get_or_create_person(
                nombre=str(raw_name or '').strip(),
                rut=self._cell(ws, row, headers, 'rut'),
                dv=self._cell(ws, row, headers, 'dv'),
                phone1=self._cell(ws, row, headers, 'telefono fijo'),
                phone2=self._cell(ws, row, headers, 'telefono movil'),
                email=self._cell(ws, row, headers, 'e mail', 'email'),
                notes=self._cell(ws, row, headers, 'obs esp'),
                counter=counter,
            )
            if parcel and person:
                self._upsert_ownership(parcel, person, OwnershipType.PRINCIPAL, counter)

    def _parse_otros_duenos(self, ws, job, sheet_result, counter: Counter):
        for row in range(2, ws.max_row + 1):
            parcela = ws.cell(row=row, column=1).value
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)

            for offset in (2, 5, 8, 11, 14, 17):
                rut = ws.cell(row=row, column=offset).value
                dv = ws.cell(row=row, column=offset + 1).value
                nombre = ws.cell(row=row, column=offset + 2).value
                if not nombre and not rut:
                    continue
                person = self._get_or_create_person(
                    nombre=str(nombre or '').strip(),
                    rut=rut,
                    dv=dv,
                    counter=counter,
                    dry_create=True,
                )
                if parcel and person:
                    self._upsert_ownership(parcel, person, OwnershipType.COPROPIETARIO, counter)

    def _parse_residentes(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'residente'])
        if not header_row:
            counter.warnings += 1
            self._issue(job, sheet_result, IssueSeverity.WARNING, ws.title, None, None, 'header_not_found', 'No se detectó encabezado en RESIDENTES')
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            estado_residente = self._cell(ws, row, headers, 'residente')
            observaciones = self._cell(ws, row, headers, 'observaciones')
            if not parcela:
                continue

            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            tipo = ResidentType.CUIDADOR if 'cuid' in str(estado_residente or '').lower() else ResidentType.RESIDENTE
            active = 'INACT' not in str(estado_residente or '').upper()
            obs = str(observaciones or '').strip()
            existing = ParcelResident.objects.filter(
                parcela=parcel,
                persona__isnull=True,
                tipo_residencia=tipo,
                observaciones=obs,
                is_deleted=False,
            ).first()
            if existing:
                if existing.is_active != active and not self.dry_run:
                    existing.is_active = active
                    existing.save(update_fields=['is_active', 'updated_at'])
                    counter.updated += 1
                else:
                    counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            ParcelResident.objects.create(
                parcela=parcel,
                persona=None,
                tipo_residencia=tipo,
                is_active=active,
                observaciones=obs,
            )
            counter.inserted += 1

    def _parse_vehiculos(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'ppu'])
        if not header_row:
            counter.errors += 1
            self._issue(job, sheet_result, IssueSeverity.ERROR, ws.title, None, None, 'header_not_found', 'No se detectó encabezado en PPU_LOGOS')
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            ppu = self._cell(ws, row, headers, 'ppu')
            if not parcela or not ppu:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            ppu_norm = ''.join(ch for ch in str(ppu).upper() if ch.isalnum())
            defaults = {
                'marca': str(self._cell(ws, row, headers, 'marca') or '').strip(),
                'tipo': str(self._cell(ws, row, headers, 'tipo') or '').strip(),
                'color': str(self._cell(ws, row, headers, 'color') or '').strip(),
                'codigo_acceso': str(self._cell(ws, row, headers, 'codigo') or '').strip(),
                'ppu': str(ppu).strip().upper(),
                'activo': True,
            }
            existing = Vehicle.objects.filter(parcela=parcel, ppu_normalizado=ppu_norm, is_deleted=False).first()
            if existing:
                changed = any(getattr(existing, k) != v for k, v in defaults.items() if v)
                if changed and not self.dry_run:
                    for key, val in defaults.items():
                        if val:
                            setattr(existing, key, val)
                    existing.save()
                    counter.updated += 1
                else:
                    counter.skipped += 1
                continue

            if self.dry_run:
                counter.inserted += 1
                continue
            Vehicle.objects.create(parcela=parcel, **defaults)
            counter.inserted += 1

    def _parse_mora_gc(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'mora cg uf', 'total pesos'])
        if not header_row:
            counter.warnings += 1
            self._issue(job, sheet_result, IssueSeverity.WARNING, ws.title, None, None, 'header_not_found', 'No se detectó encabezado en Mora GC')
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            defaults = {
                'numero_gastos_comunes': self._to_int(self._cell(ws, row, headers, 'n gastos comunes')),
                'mora_uf': self._to_decimal(self._cell(ws, row, headers, 'mora cg uf')),
                'interes_mora_uf': self._to_decimal(self._cell(ws, row, headers, 'interes mora uf')),
                'total_uf': self._to_decimal(self._cell(ws, row, headers, 'total uf')),
                'total_pesos': self._to_decimal(self._cell(ws, row, headers, 'total pesos')),
                'estado_pago': 'PENDIENTE',
            }
            duplicate = CommonExpenseDebt.objects.filter(
                parcela=parcel,
                numero_gastos_comunes=defaults['numero_gastos_comunes'],
                total_pesos=defaults['total_pesos'],
                total_uf=defaults['total_uf'],
                is_deleted=False,
            ).exists()
            if duplicate:
                counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            CommonExpenseDebt.objects.create(parcela=parcel, **defaults)
            counter.inserted += 1

    def _parse_deudas_ays(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'total deuda'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            saldo_total = self._to_decimal(self._cell(ws, row, headers, 'total deuda'))
            defaults = {
                'tipo_servicio': ServiceType.AYS,
                'numero_boletas': self._to_int(self._cell(ws, row, headers, 'boletas')),
                'monto_total': self._to_decimal(self._cell(ws, row, headers, 'a s total', 'total')),
                'convenios': self._to_decimal(self._cell(ws, row, headers, 'convenios')),
                'anticipos': self._to_decimal(self._cell(ws, row, headers, 'anticipos')),
                'saldo_total': saldo_total,
                'estado_pago': 'PENDIENTE' if saldo_total > 0 else 'PAGADO',
                'observaciones': str(self._cell(ws, row, headers, 'comentarios') or ''),
            }
            existing = ServiceDebt.objects.filter(
                parcela=parcel,
                tipo_servicio=ServiceType.AYS,
                is_deleted=False,
            ).order_by('-created_at').first()
            if existing and existing.saldo_total == defaults['saldo_total'] and existing.numero_boletas == defaults['numero_boletas']:
                counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            if existing:
                for key, value in defaults.items():
                    setattr(existing, key, value)
                existing.save()
                counter.updated += 1
            else:
                ServiceDebt.objects.create(parcela=parcel, **defaults)
                counter.inserted += 1

    def _parse_mora_convenio(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'total mora'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            saldo = self._to_decimal(self._cell(ws, row, headers, 'total mora'))
            if saldo <= 0:
                counter.skipped += 1
                continue

            defaults = {
                'empresa': str(self._cell(ws, row, headers, 'cobranza') or ''),
                'tipo': 'MORA_CONVENIO',
                'detalle': f"GC: {self._to_int(self._cell(ws, row, headers, 'n gc'))}",
                'saldo_monto': saldo,
                'estado_pago': 'PENDIENTE',
            }
            duplicate = PaymentAgreement.objects.filter(
                parcela=parcel,
                empresa=defaults['empresa'],
                tipo=defaults['tipo'],
                detalle=defaults['detalle'],
                saldo_monto=defaults['saldo_monto'],
                is_deleted=False,
            ).exists()
            if duplicate:
                counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            PaymentAgreement.objects.create(parcela=parcel, **defaults)
            counter.inserted += 1

    def _parse_multas(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'empresa', 'saldo monto'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            saldo = self._to_decimal(self._cell(ws, row, headers, 'saldo monto'))
            defaults = {
                'empresa': str(self._cell(ws, row, headers, 'empresa') or ''),
                'tipo': str(self._cell(ws, row, headers, 'tipo') or ''),
                'fecha_emision': self._to_date(self._cell(ws, row, headers, 'emision')),
                'fecha_vencimiento': self._to_date(self._cell(ws, row, headers, 'vencimiento')),
                'detalle': str(self._cell(ws, row, headers, 'detalle') or ''),
                'saldo_monto': saldo,
                'estado_pago': 'PENDIENTE' if saldo > 0 else 'PAGADO',
            }
            duplicate = UnpaidFine.objects.filter(
                parcela=parcel,
                empresa=defaults['empresa'],
                tipo=defaults['tipo'],
                fecha_vencimiento=defaults['fecha_vencimiento'],
                saldo_monto=defaults['saldo_monto'],
                is_deleted=False,
            ).exists()
            if duplicate:
                counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            UnpaidFine.objects.create(parcela=parcel, **defaults)
            counter.inserted += 1

    def _parse_cortes(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['cliente', 'estado'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'cliente')
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            estado = str(self._cell(ws, row, headers, 'estado') or '').strip()
            te1 = str(self._cell(ws, row, headers, 'te1 vencimiento') or '').strip()
            corte_luz = str(self._cell(ws, row, headers, 'corte luz') or '').strip()
            corte_ap = str(self._cell(ws, row, headers, 'corte ap') or '').strip()

            tipo = CutType.AYS
            if corte_luz and not corte_ap:
                tipo = CutType.LUZ
            elif corte_ap and not corte_luz:
                tipo = CutType.AGUA

            defaults = {
                'tipo_corte': tipo,
                'estado': estado,
                'motivo': te1,
                'fecha': self._to_date(self._cell(ws, row, headers, 'fecha')),
                'activo': True,
            }
            existing = ServiceCut.objects.filter(
                parcela=parcel,
                tipo_corte=defaults['tipo_corte'],
                fecha=defaults['fecha'],
                motivo=defaults['motivo'],
                is_deleted=False,
            ).first()
            if existing:
                changed = existing.estado != defaults['estado'] or existing.activo != defaults['activo']
                if changed and not self.dry_run:
                    existing.estado = defaults['estado']
                    existing.activo = defaults['activo']
                    existing.save(update_fields=['estado', 'activo', 'updated_at'])
                    counter.updated += 1
                else:
                    counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            ServiceCut.objects.create(parcela=parcel, **defaults)
            counter.inserted += 1

    def _parse_historico_ays(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'solicitante', 'descripcion'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            if not parcela:
                continue
            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            defaults = {
                'numero_orden': str(self._cell(ws, row, headers, 'orden') or ''),
                'solicitante': str(self._cell(ws, row, headers, 'solicitante') or ''),
                'resultado': str(self._cell(ws, row, headers, 'realizado') or ''),
                'descripcion': str(self._cell(ws, row, headers, 'descripcion') or ''),
                'fecha_ingreso': self._to_date(self._cell(ws, row, headers, 'fecha ingreso')),
                'fecha_ejecucion': self._to_date(self._cell(ws, row, headers, 'fecha ejecucion')),
                'ejecutante': str(self._cell(ws, row, headers, 'ejecutante') or ''),
                'lugar_corte_reposicion': str(self._cell(ws, row, headers, 'lugar de corte') or ''),
                'observaciones': str(self._cell(ws, row, headers, 'obvervaciones') or ''),
            }
            duplicate = ServiceHistory.objects.filter(
                parcela=parcel,
                numero_orden=defaults['numero_orden'],
                descripcion=defaults['descripcion'],
                fecha_ingreso=defaults['fecha_ingreso'],
                is_deleted=False,
            ).exists()
            if duplicate:
                counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue
            ServiceHistory.objects.create(parcela=parcel, **defaults)
            counter.inserted += 1

    def _parse_anotaciones(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela', 'fecha', 'anotacion'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela')
            texto = self._cell(ws, row, headers, 'anotacion')
            if not parcela or not texto:
                continue

            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue
            event_date = self._to_date(self._cell(ws, row, headers, 'fecha'))
            normalized_text = str(texto).strip()
            duplicate = AdministrativeNote.objects.filter(
                parcela=parcel,
                texto=normalized_text,
                fecha_evento=event_date,
                is_deleted=False,
            ).exists()
            if duplicate:
                counter.skipped += 1
                continue
            if self.dry_run:
                counter.inserted += 1
                continue

            AdministrativeNote.objects.create(
                parcela=parcel,
                tipo=NoteType.ADMINISTRATIVA,
                texto=normalized_text,
                fecha_evento=event_date,
            )
            counter.inserted += 1

    def _parse_obras(self, ws, job, sheet_result, counter: Counter):
        header_row, headers = self._find_header(ws, ['parcela n', 'cortafuego', 'limpieza'])
        if not header_row:
            counter.warnings += 1
            return

        for row in range(header_row + 1, ws.max_row + 1):
            parcela = self._cell(ws, row, headers, 'parcela n')
            if not parcela:
                continue

            counter.rows_read += 1
            parcel = self._upsert_parcel(parcela, counter, job, sheet_result, row)
            if not parcel:
                continue

            defaults = {
                'deshabitada': str(self._cell(ws, row, headers, 'deshabitada') or ''),
                'cercada': str(self._cell(ws, row, headers, 'cercada') or ''),
                'sucia': str(self._cell(ws, row, headers, 'sucia') or ''),
                'casas': str(self._cell(ws, row, headers, 'casas') or ''),
                'otra_construccion': str(self._cell(ws, row, headers, 'otra const') or ''),
                'cumplen': str(self._cell(ws, row, headers, 'cumplen') or ''),
                'cortafuego': str(self._cell(ws, row, headers, 'cortafuego') or ''),
                'limpieza': str(self._cell(ws, row, headers, 'limpieza') or ''),
                'foco_incendio': str(self._cell(ws, row, headers, 'foco incend') or ''),
                'atributo_kpi': self._to_decimal(self._cell(ws, row, headers, 'atributo kpi'), default=None),
                'kpi': str(self._cell(ws, row, headers, 'kpi') or ''),
                'estado_actual': str(self._cell(ws, row, headers, 'estado actual') or ''),
                'rol_sii': str(self._cell(ws, row, headers, 'rol') or ''),
                'certificado_obras': str(self._cell(ws, row, headers, 'certificado obras') or ''),
                'permiso_dom': str(self._cell(ws, row, headers, 'permiso dom') or ''),
            }

            if self.dry_run:
                counter.inserted += 1
                continue

            _, created = ParcelWorkStatus.objects.update_or_create(parcela=parcel, defaults=defaults)
            if created:
                counter.inserted += 1
            else:
                counter.updated += 1

